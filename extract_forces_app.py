"""
Streamlit app — Element Forces Analyser (Tables Only)
Extracts min/max F11 (+ M22) and min/max F22 (+ M11) from an Excel shell-forces report.
Also analyses M11/M22 extremes, identifies their Area Shells, and extracts F22/F11 from those areas.

Run with:
    streamlit run extract_forces_app.py
"""

import openpyxl
import pandas as pd
import streamlit as st


# ─────────────────────────────────────────────
# Core extraction logic
# ─────────────────────────────────────────────

def extract_all(filepath) -> tuple[dict, dict, pd.DataFrame]:
    wb = openpyxl.load_workbook(filepath)
    ws = wb["Element Forces - Area Shells"]

    headers = [cell.value for cell in ws[2]]
    col = {name: headers.index(name) for name in ("Area", "F11", "F22", "M11", "M22")}

    res = {
        "F11_max": {"F11": None, "M22": None, "row": None},
        "F11_min": {"F11": None, "M22": None, "row": None},
        "F22_max": {"F22": None, "M11": None, "row": None},
        "F22_min": {"F22": None, "M11": None, "row": None},
    }

    m_res = {
        "M11_max": {"M11": None, "F22": None, "Area": None, "row": None},
        "M11_min": {"M11": None, "F22": None, "Area": None, "row": None},
        "M22_max": {"M22": None, "F11": None, "Area": None, "row": None},
        "M22_min": {"M22": None, "F11": None, "Area": None, "row": None},
    }

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        area = row[col["Area"]]
        f11  = row[col["F11"]]
        f22  = row[col["F22"]]
        m11  = row[col["M11"]]
        m22  = row[col["M22"]]

        if not all(isinstance(v, (int, float)) for v in (f11, f22, m11, m22)):
            continue

        rows.append({"row": row_idx, "Area": area, "F11": f11, "F22": f22, "M11": m11, "M22": m22})

        if res["F11_max"]["F11"] is None or f11 > res["F11_max"]["F11"]:
            res["F11_max"] = {"F11": f11, "M22": m22, "row": row_idx}
        if res["F11_min"]["F11"] is None or f11 < res["F11_min"]["F11"]:
            res["F11_min"] = {"F11": f11, "M22": m22, "row": row_idx}
        if res["F22_max"]["F22"] is None or f22 > res["F22_max"]["F22"]:
            res["F22_max"] = {"F22": f22, "M11": m11, "row": row_idx}
        if res["F22_min"]["F22"] is None or f22 < res["F22_min"]["F22"]:
            res["F22_min"] = {"F22": f22, "M11": m11, "row": row_idx}

        if m_res["M11_max"]["M11"] is None or m11 > m_res["M11_max"]["M11"]:
            m_res["M11_max"] = {"M11": m11, "F22": f22, "Area": area, "row": row_idx}
        if m_res["M11_min"]["M11"] is None or m11 < m_res["M11_min"]["M11"]:
            m_res["M11_min"] = {"M11": m11, "F22": f22, "Area": area, "row": row_idx}
        if m_res["M22_max"]["M22"] is None or m22 > m_res["M22_max"]["M22"]:
            m_res["M22_max"] = {"M22": m22, "F11": f11, "Area": area, "row": row_idx}
        if m_res["M22_min"]["M22"] is None or m22 < m_res["M22_min"]["M22"]:
            m_res["M22_min"] = {"M22": m22, "F11": f11, "Area": area, "row": row_idx}

    return res, m_res, df


# ─────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────

st.set_page_config(page_title="Element Forces Analyser", page_icon="🏗️", layout="wide")

st.title("🏗️ Element Forces Analyser")
st.markdown("Upload your **Element Forces – Area Shells** Excel file to extract extreme force and moment values.")

uploaded = st.file_uploader("Upload Excel file (.xlsx)", type=["xlsx"])

if uploaded is None:
    st.info("👆 Upload an `.xlsx` file to get started.")
    st.stop()

with st.spinner("Reading and analysing…"):
    res, m_res, df = extract_all(uploaded)

st.success(f"✅ Analysed **{len(df):,}** valid data rows.")


# ════════════════════════════════════════════
# SECTION 1 — F11 & F22 Extremes
# ════════════════════════════════════════════
st.markdown("---")
st.header("📐 Section 1 — F11 & F22 Extremes")

c1, c2, c3, c4 = st.columns(4)
with c1:
    st.metric("Max F11 (KN/m)", f"{res['F11_max']['F11']:.4f}",
              f"M22 = {res['F11_max']['M22']:.4f} KN-m/m")
with c2:
    st.metric("Min F11 (KN/m)", f"{res['F11_min']['F11']:.4f}",
              f"M22 = {res['F11_min']['M22']:.4f} KN-m/m", delta_color="inverse")
with c3:
    st.metric("Max F22 (KN/m)", f"{res['F22_max']['F22']:.4f}",
              f"M11 = {res['F22_max']['M11']:.4f} KN-m/m")
with c4:
    st.metric("Min F22 (KN/m)", f"{res['F22_min']['F22']:.4f}",
              f"M11 = {res['F22_min']['M11']:.4f} KN-m/m", delta_color="inverse")

st.markdown(" ")
ca, cb = st.columns(2)

with ca:
    st.subheader("F11 → M22")
    st.dataframe(pd.DataFrame([
        {"Extreme": "Max F11", "F11 (KN/m)": res["F11_max"]["F11"],
         "M22 (KN-m/m)": res["F11_max"]["M22"], "Excel Row": res["F11_max"]["row"]},
        {"Extreme": "Min F11", "F11 (KN/m)": res["F11_min"]["F11"],
         "M22 (KN-m/m)": res["F11_min"]["M22"], "Excel Row": res["F11_min"]["row"]},
    ]), use_container_width=True, hide_index=True)

with cb:
    st.subheader("F22 → M11")
    st.dataframe(pd.DataFrame([
        {"Extreme": "Max F22", "F22 (KN/m)": res["F22_max"]["F22"],
         "M11 (KN-m/m)": res["F22_max"]["M11"], "Excel Row": res["F22_max"]["row"]},
        {"Extreme": "Min F22", "F22 (KN/m)": res["F22_min"]["F22"],
         "M11 (KN-m/m)": res["F22_min"]["M11"], "Excel Row": res["F22_min"]["row"]},
    ]), use_container_width=True, hide_index=True)


# ════════════════════════════════════════════
# SECTION 2 — M11 Extremes + paired F22
# ════════════════════════════════════════════
st.markdown("---")
st.header("🔎 Section 2 — M11 Extremes")
st.markdown("The **min and max M11** values and their corresponding **F22** at the same row.")

s2c1, s2c2, s2c3, s2c4 = st.columns(4)
with s2c1:
    st.metric("Max M11 (KN-m/m)", f"{m_res['M11_max']['M11']:.4f}",
              f"F22 = {m_res['M11_max']['F22']:.4f} KN/m")
with s2c2:
    st.metric("Min M11 (KN-m/m)", f"{m_res['M11_min']['M11']:.4f}",
              f"F22 = {m_res['M11_min']['F22']:.4f} KN/m", delta_color="inverse")
with s2c3:
    st.metric("Area (Max M11)", str(m_res["M11_max"]["Area"]))
with s2c4:
    st.metric("Area (Min M11)", str(m_res["M11_min"]["Area"]))

st.markdown(" ")
st.subheader("M11 → F22")
st.dataframe(pd.DataFrame([
    {"Extreme": "Max M11", "M11 (KN-m/m)": m_res["M11_max"]["M11"],
     "F22 (KN/m)": m_res["M11_max"]["F22"], "Area": m_res["M11_max"]["Area"],
     "Excel Row": m_res["M11_max"]["row"]},
    {"Extreme": "Min M11", "M11 (KN-m/m)": m_res["M11_min"]["M11"],
     "F22 (KN/m)": m_res["M11_min"]["F22"], "Area": m_res["M11_min"]["Area"],
     "Excel Row": m_res["M11_min"]["row"]},
]), use_container_width=True, hide_index=True)


# ════════════════════════════════════════════
# SECTION 3 — M22 Extremes + paired F11
# ════════════════════════════════════════════
st.markdown("---")
st.header("🔎 Section 3 — M22 Extremes")
st.markdown("The **min and max M22** values and their corresponding **F11** at the same row.")

s3c1, s3c2, s3c3, s3c4 = st.columns(4)
with s3c1:
    st.metric("Max M22 (KN-m/m)", f"{m_res['M22_max']['M22']:.4f}",
              f"F11 = {m_res['M22_max']['F11']:.4f} KN/m")
with s3c2:
    st.metric("Min M22 (KN-m/m)", f"{m_res['M22_min']['M22']:.4f}",
              f"F11 = {m_res['M22_min']['F11']:.4f} KN/m", delta_color="inverse")
with s3c3:
    st.metric("Area (Max M22)", str(m_res["M22_max"]["Area"]))
with s3c4:
    st.metric("Area (Min M22)", str(m_res["M22_min"]["Area"]))

st.markdown(" ")
st.subheader("M22 → F11")
st.dataframe(pd.DataFrame([
    {"Extreme": "Max M22", "M22 (KN-m/m)": m_res["M22_max"]["M22"],
     "F11 (KN/m)": m_res["M22_max"]["F11"], "Area": m_res["M22_max"]["Area"],
     "Excel Row": m_res["M22_max"]["row"]},
    {"Extreme": "Min M22", "M22 (KN-m/m)": m_res["M22_min"]["M22"],
     "F11 (KN/m)": m_res["M22_min"]["F11"], "Area": m_res["M22_min"]["Area"],
     "Excel Row": m_res["M22_min"]["row"]},
]), use_container_width=True, hide_index=True)


# ════════════════════════════════════════════
# Raw Data
# ════════════════════════════════════════════
st.markdown("---")
with st.expander("📋 View Raw Data Table"):
    st.dataframe(df, use_container_width=True, hide_index=True)
